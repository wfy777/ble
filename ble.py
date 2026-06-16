#!/usr/bin/env python3
# Copyright (C) 2026 F. Y. Wu
# SPDX-License-Identifier: GPL-3.0-or-later
"""
Desktop UI for BLE scan, connect, GATT discovery, hex write, and notifications.

Requires: pip install bleak openpyxl
Uses tkinter (stdlib). All BLE runs on a dedicated asyncio thread.
"""

from __future__ import annotations

import asyncio
import concurrent.futures
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from dataclasses import dataclass
from typing import Any, Awaitable, Callable, Iterable, Optional, Sequence, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    from bleak import BleakClient, BleakScanner
    from bleak.backends.characteristic import BleakGATTCharacteristic
except ImportError as e:
    raise SystemExit("Install bleak: pip install bleak") from e

try:
    from ble_industrial_tool import fmt_hex, parse_hex
except ImportError:
    # Allow running as single script if copied elsewhere
    def parse_hex(s: str) -> bytes:
        cleaned = "".join(s.split())
        if len(cleaned) % 2 != 0:
            raise ValueError("Hex string must have an even number of digits")
        return bytes.fromhex(cleaned)

    def fmt_hex(data: bytes) -> str:
        return data.hex(" ").upper()


def crc16_modbus(data: bytes) -> int:
    crc = 0xFFFF
    for byte in data:
        crc ^= byte
        for _ in range(8):
            if crc & 1:
                crc = (crc >> 1) ^ 0xA001
            else:
                crc >>= 1
    return crc & 0xFFFF


def parse_int_for_frame(s: str) -> int:
    t = str(s).strip()
    if not t:
        raise ValueError("empty value")
    if t.lower().startswith("0x"):
        return int(t, 16)
    return int(float(t))


def first_int_from_text_as_byte(s: str) -> int:
    """取单元格中第一个整数值，作为单字节 (0–255)。"""
    m = re.search(r"-?\d+", str(s))
    if not m:
        raise ValueError(f"No integer found in cell text: {s!r}")
    return int(m.group(0)) & 0xFF


def _is_address_anchor_cell(v: Any) -> bool:
    if v is None:
        return False
    t = str(v).strip().lower()
    return t in ("address", "地址")


def parse_worksheet_seven_from_address_anchor(ws: Any) -> tuple[int, int, list[str], list[list[str]]]:
    """Return (anchor_row_1based, anchor_col_1based, 7 headers, data rows each 7 str) from one worksheet."""
    anchor_r: Optional[int] = None
    anchor_c: Optional[int] = None
    for row in ws.iter_rows():
        for cell in row:
            if _is_address_anchor_cell(cell.value):
                anchor_r, anchor_c = cell.row, cell.column
                break
        if anchor_r is not None:
            break
    if anchor_r is None or anchor_c is None:
        raise ValueError(
            f'No anchor cell on sheet {ws.title!r}: need a cell whose text is exactly "address" or "地址".'
        )
    headers: list[str] = []
    for j in range(7):
        cell = ws.cell(row=anchor_r, column=anchor_c + j)
        val = cell.value
        headers.append("" if val is None else str(val))
    data: list[list[str]] = []
    r = anchor_r + 1
    while r < anchor_r + 5000:
        rowvals: list[str] = []
        blank = True
        for j in range(7):
            cell = ws.cell(row=r, column=anchor_c + j)
            val = cell.value
            s = "" if val is None else str(val)
            if str(s).strip():
                blank = False
            rowvals.append(s)
        if blank:
            break
        data.append(rowvals)
        r += 1
    return anchor_r, anchor_c, headers, data


def excel_sheet_names(path: str) -> list[str]:
    if openpyxl is None:
        raise RuntimeError("Install openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def load_excel_seven_from_address_anchor(
    path: str,
    sheet: Optional[str] = None,
) -> tuple[int, int, list[str], list[list[str]], str]:
    """Return (anchor_row, anchor_col, headers, data rows, sheet_name)."""
    if openpyxl is None:
        raise RuntimeError("Install openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet not found: {sheet!r}")
            ws = wb[sheet]
        else:
            ws = wb.active
        ar, ac, headers, data = parse_worksheet_seven_from_address_anchor(ws)
        return ar, ac, headers, data, ws.title
    finally:
        wb.close()


def function_attribute_col_index(headers: Sequence[str]) -> int:
    """0-based column index within the 7 Excel columns."""
    for i, h in enumerate(headers[:7]):
        t = str(h).strip().lower()
        if t in ("功能属性", "functional attribute", "function attribute", "func attr"):
            return i
        if "function" in t and "attr" in t.replace(" ", ""):
            return i
    raise ValueError('No "functional attribute" / "功能属性" column in the anchored header row.')


def _address_column_is_one(addr_cell: str) -> bool:
    s = str(addr_cell).strip()
    if s == "1":
        return True
    try:
        return int(float(s)) == 1
    except (TypeError, ValueError):
        return False


def _wrap_protocol_body(body: bytes | bytearray) -> bytes:
    crc = crc16_modbus(bytes(body))
    return bytes([0x55, 0xAA, *body, crc & 0xFF, (crc >> 8) & 0xFF, 0x0D, 0x0A])


def feature_byte_from_excel_rows(
    rows: Sequence[Sequence[str]],
    func_col_idx: int,
    *,
    addr_col_idx: int = 0,
) -> int:
    """Feature byte from the row whose address column equals 1."""
    for row in rows:
        vals = tuple(str(x) for x in (list(row) + [""] * 7)[:7])
        if addr_col_idx >= len(vals):
            continue
        if _address_column_is_one(vals[addr_col_idx]):
            if func_col_idx >= len(vals):
                raise ValueError("Functional attribute column index out of range.")
            return first_int_from_text_as_byte(str(vals[func_col_idx]))
    raise ValueError('No data row where the **address** column (first column of the 7) equals 1.')


def build_write_register_frame(feat_b: int, addr: int, data: int) -> bytes:
    """Modbus-like write (0x10): one 16-bit register at addr."""
    addr &= 0xFFFF
    data &= 0xFFFF
    body = bytearray([feat_b & 0xFF, 0x10])
    body.extend([(addr >> 8) & 0xFF, addr & 0xFF, 0x00, 0x01, (data >> 8) & 0xFF, data & 0xFF])
    return _wrap_protocol_body(body)


def build_read_register_frame(feat_b: int, addr: int) -> bytes:
    """Modbus-like read holding register (0x03): one register at addr."""
    addr &= 0xFFFF
    body = bytearray([feat_b & 0xFF, 0x03])
    body.extend([(addr >> 8) & 0xFF, addr & 0xFF, 0x00, 0x01])
    return _wrap_protocol_body(body)


def _extract_protocol_body(raw: bytes) -> bytes:
    """Return validated body bytes (between 55 AA and CRC, CRC excluded)."""
    data = raw
    idx = data.find(b"\x55\xAA")
    if idx >= 0:
        data = data[idx:]
    if len(data) < 9 or data[:2] != b"\x55\xAA" or data[-2:] != b"\x0D\x0A":
        raise ValueError("invalid response frame")
    body_plus_crc = data[2:-2]
    if len(body_plus_crc) < 3:
        raise ValueError("response too short")
    body = body_plus_crc[:-2]
    crc_got = body_plus_crc[-2:]
    crc_exp = crc16_modbus(body)
    if crc_got[0] != (crc_exp & 0xFF) or crc_got[1] != ((crc_exp >> 8) & 0xFF):
        raise ValueError("response CRC mismatch")
    return bytes(body)


def format_read_data_display(data: bytes) -> str:
    """Format read payload as 'dec (0xHEX)' per 16-bit register."""
    if not data:
        raise ValueError("empty read data")
    if len(data) % 2 != 0:
        raise ValueError("read data length is not a multiple of 2")
    parts: list[str] = []
    for i in range(0, len(data), 2):
        v = ((data[i] & 0xFF) << 8) | (data[i + 1] & 0xFF)
        parts.append(f"{v} (0x{v:04X})")
    return ", ".join(parts)


def format_register_display(value: int) -> str:
    value &= 0xFFFF
    return f"{value} (0x{value:04X})"


def parse_register_value_text(text: str) -> int:
    """Accept plain decimal/hex or '1234 (0x04D2)' from the read display."""
    t = str(text).strip()
    if not t:
        raise ValueError("empty value")
    if "," in t:
        t = t.split(",", 1)[0].strip()
    if "(" in t:
        t = t.split("(", 1)[0].strip()
    return parse_int_for_frame(t) & 0xFFFF


def parse_read_register_response(raw: bytes) -> bytes:
    """
    Parse read response frame:
    55 AA feat 03 address(2) length(2) data(length*2) crc(2) 0D 0A
    """
    body = _extract_protocol_body(raw)
    if len(body) < 6:
        raise ValueError("response body too short")
    func = body[1]
    if func & 0x80:
        raise ValueError("device returned modbus exception")
    if func != 0x03:
        raise ValueError(f"unexpected function code 0x{func:02X}")
    reg_count = ((body[4] & 0xFF) << 8) | (body[5] & 0xFF)
    if reg_count <= 0:
        raise ValueError("invalid read length")
    data_len = reg_count * 2
    if len(body) < 6 + data_len:
        raise ValueError("response missing read data")
    return bytes(body[6 : 6 + data_len])


def build_protocol_frame_bytes(
    rows: Sequence[Tuple[str, Tuple[str, ...]]],
    *,
    func_col_idx: int,
    addr_col_idx: int = 0,
) -> bytes:
    """
    rows: top-to-bottom (fill_column_text, tuple of 7 Excel cell strings).
    Pick the row where **address column** (default first column) equals 1; read that row's
    **functional attribute** cell for the feature byte (first integer in cell, one byte).
    Then 0x10, then for each row with non-empty fill: big-endian 16-bit address, 00 01,
    big-endian 16-bit data, CRC-16-Modbus over all bytes after 55 AA (low CRC byte first).
    """
    excel_rows = [tuple(str(vals[i]) if i < len(vals) else "" for i in range(7)) for _fill, vals in rows]
    feat_b = feature_byte_from_excel_rows(excel_rows, func_col_idx, addr_col_idx=addr_col_idx)

    body = bytearray([feat_b, 0x10])
    for fill, vals in rows:
        fill_s = str(fill).strip()
        if not fill_s:
            continue
        addr_s = str(vals[addr_col_idx]).strip() if addr_col_idx < len(vals) else ""
        if not addr_s:
            raise ValueError(f"Fill is set but address column is empty in row with fill={fill_s!r}")
        addr_v = parse_int_for_frame(addr_s) & 0xFFFF
        data_v = parse_int_for_frame(fill_s) & 0xFFFF
        body.extend([(addr_v >> 8) & 0xFF, addr_v & 0xFF, 0x00, 0x01, (data_v >> 8) & 0xFF, data_v & 0xFF])

    return _wrap_protocol_body(body)


def _norm_uuid(u: Any) -> str:
    return str(u).lower().replace("-", "")


def _prop_set(props: Optional[Iterable[Any]]) -> set[str]:
    """Normalize GATT property flags across Bleak backends."""
    out: set[str] = set()
    for p in props or []:
        s = str(p).strip().lower().replace("_", "-")
        if s in ("write", "write-without-response", "notify", "indicate"):
            out.add(s)
        elif "write" in s and "without" in s:
            out.add("write-without-response")
        elif s == "read":
            out.add("read")
        elif "notify" in s:
            out.add("notify")
        elif "indicate" in s:
            out.add("indicate")
        elif s.startswith("write"):
            out.add("write")
    return out


def _write_flags(ps: set[str]) -> tuple[bool, bool]:
    """Returns (can_write_with_response, can_write_without_response)."""
    wr = "write" in ps
    wo = "write-without-response" in ps
    return wr, wo


def _infer_write_without_response(ps: set[str]) -> bool:
    wr, wo = _write_flags(ps)
    if wo and not wr:
        return True
    if wr and not wo:
        return False
    if wr and wo:
        return False
    return False


def _can_notify(ps: set[str]) -> bool:
    return "notify" in ps or "indicate" in ps


# Nordic UART Service (common on industrial BLE–serial bridges)
_NUS_SVC = _norm_uuid("6E400001-B5A3-F393-E0A9-E50E24DCCA9E")
_NUS_RX = _norm_uuid("6E400002-B5A3-F393-E0A9-E50E24DCCA9E")
_NUS_TX = _norm_uuid("6E400003-B5A3-F393-E0A9-E50E24DCCA9E")


@dataclass
class AutoUuidResult:
    write_uuid: Optional[str]
    notify_uuid: Optional[str]
    write_without_response: bool
    note: str


def pick_data_characteristics(services: Any) -> AutoUuidResult:
    """
    Guess which GATT characteristics carry app data (hex send / notify receive).

    Order: Nordic UART → one write + one notify in the same service → globally
    unique write & notify → first service that has both → any write + any notify.
    """
    writes: dict[str, set[str]] = {}
    notify_order: list[str] = []
    notify_seen: set[str] = set()

    for svc in services:
        su = _norm_uuid(svc.uuid)
        if su == _NUS_SVC:
            rx: Optional[str] = None
            tx: Optional[str] = None
            rx_props: set[str] = set()
            for ch in svc.characteristics:
                ps = _prop_set(ch.properties)
                cu = str(ch.uuid)
                cn = _norm_uuid(cu)
                if cn == _NUS_RX:
                    rx, rx_props = cu, ps
                elif cn == _NUS_TX:
                    tx = cu
            if rx and tx:
                return AutoUuidResult(
                    rx,
                    tx,
                    _infer_write_without_response(rx_props),
                    "Detected Nordic UART Service (NUS): standard RX/TX characteristics.",
                )

    per_service: list[tuple[str, dict[str, set[str]], list[str]]] = []
    for svc in services:
        su = str(svc.uuid)
        w_sub: dict[str, set[str]] = {}
        n_sub: list[str] = []
        n_seen: set[str] = set()
        for ch in svc.characteristics:
            ps = _prop_set(ch.properties)
            cu = str(ch.uuid)
            wr, wo = _write_flags(ps)
            if wr or wo:
                w_sub[cu] = ps
                writes[cu] = ps
            if _can_notify(ps) and cu not in n_seen:
                n_seen.add(cu)
                n_sub.append(cu)
                if cu not in notify_seen:
                    notify_seen.add(cu)
                    notify_order.append(cu)
        per_service.append((su, w_sub, n_sub))

    for su, w_sub, n_sub in sorted(per_service, key=lambda t: _norm_uuid(t[0])):
        if len(w_sub) == 1 and len(n_sub) == 1:
            w_uuid, w_ps = next(iter(w_sub.items()))
            return AutoUuidResult(
                w_uuid,
                n_sub[0],
                _infer_write_without_response(w_ps),
                f"Single write + single notify in service {su}.",
            )

    all_writes = list(writes.items())
    if len(all_writes) == 1 and len(notify_order) == 1:
        w_uuid, w_ps = all_writes[0]
        return AutoUuidResult(
            w_uuid,
            notify_order[0],
            _infer_write_without_response(w_ps),
            "Only one writable and one notifiable characteristic on the device.",
        )

    for su, w_sub, n_sub in sorted(per_service, key=lambda t: _norm_uuid(t[0])):
        if w_sub and n_sub:
            w_uuid, w_ps = sorted(w_sub.items(), key=lambda it: _norm_uuid(it[0]))[0]
            n_uuid = sorted(n_sub, key=_norm_uuid)[0]
            return AutoUuidResult(
                w_uuid,
                n_uuid,
                _infer_write_without_response(w_ps),
                f"Best guess using service {su} (first service that has both write and notify). "
                "If traffic fails, use Discover GATT and set UUIDs manually.",
            )

    if all_writes and notify_order:
        w_uuid, w_ps = sorted(all_writes, key=lambda it: _norm_uuid(it[0]))[0]
        n_uuid = sorted(notify_order, key=_norm_uuid)[0]
        return AutoUuidResult(
            w_uuid,
            n_uuid,
            _infer_write_without_response(w_ps),
            "Best guess: first writable and first notifiable characteristic (different services possible). "
            "Verify with your device manual if this does not work.",
        )

    if all_writes:
        w_uuid, w_ps = sorted(all_writes, key=lambda it: _norm_uuid(it[0]))[0]
        return AutoUuidResult(
            w_uuid,
            None,
            _infer_write_without_response(w_ps),
            "Found writable characteristic(s) but none with notify/indicate. Receive may not be available.",
        )

    if notify_order:
        n_uuid = sorted(notify_order, key=_norm_uuid)[0]
        return AutoUuidResult(
            None,
            n_uuid,
            False,
            "Found notifiable characteristic(s) but none writable. Send may not be available.",
        )

    return AutoUuidResult(None, None, False, "No writable or notifiable characteristics found.")


class BleHub:
    """Owns one event loop thread and at most one BleakClient."""

    def __init__(self) -> None:
        self._loop: Optional[asyncio.AbstractEventLoop] = None
        self._thread: Optional[threading.Thread] = None
        self._client: Optional[BleakClient] = None
        self._notify_uuid: Optional[str] = None
        self._rx_cb: Optional[Callable[[bytes], None]] = None
        self._pending_rx: Optional[asyncio.Future[bytes]] = None
        self._started = threading.Event()

    def start(self) -> None:
        if self._thread and self._thread.is_alive():
            return

        def runner() -> None:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            self._loop = loop
            self._started.set()
            loop.run_forever()

        self._started.clear()
        self._thread = threading.Thread(target=runner, name="ble-async", daemon=True)
        self._thread.start()
        self._started.wait(timeout=5.0)
        if self._loop is None:
            raise RuntimeError("BLE event loop failed to start")

    def stop(self) -> None:
        if self._loop is None:
            return
        fut = asyncio.run_coroutine_threadsafe(self._shutdown(), self._loop)
        try:
            fut.result(timeout=15.0)
        except (concurrent.futures.TimeoutError, Exception):
            pass
        self._loop.call_soon_threadsafe(self._loop.stop)
        if self._thread:
            self._thread.join(timeout=5.0)
        self._loop = None
        self._thread = None

    async def _shutdown(self) -> None:
        await self.disconnect()

    def submit(
        self,
        coro: Awaitable[Any],
        *,
        on_done: Optional[Callable[[concurrent.futures.Future[Any]], None]] = None,
    ) -> concurrent.futures.Future[Any]:
        if self._loop is None:
            raise RuntimeError("BLE hub not started")
        fut = asyncio.run_coroutine_threadsafe(coro, self._loop)

        def _wrap(f: concurrent.futures.Future[Any]) -> None:
            if on_done:
                on_done(f)

        fut.add_done_callback(_wrap)
        return fut

    @property
    def connected(self) -> bool:
        c = self._client
        return bool(c and c.is_connected)

    def set_rx_callback(self, cb: Optional[Callable[[bytes], None]]) -> None:
        self._rx_cb = cb

    def _notify_handler(self, _char: BleakGATTCharacteristic, data: bytearray) -> None:
        b = bytes(data)
        pending = self._pending_rx
        if pending is not None and not pending.done():
            pending.set_result(b)
        cb = self._rx_cb
        if cb:
            cb(b)

    async def scan(self, timeout: float) -> list[tuple[str, Optional[str], int]]:
        # Bleak >= ~0.19: RSSI lives on AdvertisementData; use return_adv=True.
        # Older Bleak: BLEDevice.rssi may exist; no return_adv kwarg.
        rows: list[tuple[str, Optional[str], int]] = []
        try:
            found = await BleakScanner.discover(timeout=timeout, return_adv=True)
        except TypeError:
            devices = await BleakScanner.discover(timeout=timeout)
            for d in devices:
                rssi = getattr(d, "rssi", None)
                rows.append((d.address, d.name, int(rssi) if rssi is not None else 0))
        else:
            for _addr, (d, adv) in found.items():
                rssi = getattr(adv, "rssi", None)
                rows.append((d.address, d.name, int(rssi) if rssi is not None else 0))
        rows.sort(key=lambda r: (r[1] or "", r[0]))
        return rows

    async def connect(self, address: str) -> None:
        await self.disconnect()
        client = BleakClient(address)
        await client.connect()
        self._client = client

    async def disconnect(self) -> None:
        client = self._client
        if client is None:
            return
        if self._notify_uuid:
            try:
                if client.is_connected:
                    await client.stop_notify(self._notify_uuid)
            except Exception:
                pass
            self._notify_uuid = None
        try:
            if client.is_connected:
                await client.disconnect()
        except Exception:
            pass
        self._client = None

    async def discover_lines(self) -> list[str]:
        client = self._client
        if not client or not client.is_connected:
            raise RuntimeError("Not connected")
        lines: list[str] = []
        for svc in client.services:
            lines.append(f"Service {svc.uuid}")
            for ch in svc.characteristics:
                props = ",".join(ch.properties) if ch.properties else ""
                lines.append(
                    f"  Char {ch.uuid}  [{props}]  max={ch.max_write_without_response_size}"
                )
        return lines

    async def auto_pick_uuids(self) -> AutoUuidResult:
        client = self._client
        if not client or not client.is_connected:
            raise RuntimeError("Not connected")
        return pick_data_characteristics(client.services)

    async def write(self, char_uuid: str, data: bytes, *, response: bool) -> None:
        client = self._client
        if not client or not client.is_connected:
            raise RuntimeError("Not connected")
        await client.write_gatt_char(char_uuid, data, response=response)

    async def transact(
        self,
        write_uuid: str,
        notify_uuid: str,
        payload: bytes,
        *,
        response: bool,
        timeout: float = 3.0,
    ) -> bytes:
        """Write payload and wait for the next notify response."""
        await self.start_notify(notify_uuid)
        loop = asyncio.get_running_loop()
        fut: asyncio.Future[bytes] = loop.create_future()
        self._pending_rx = fut
        try:
            await self.write(write_uuid, payload, response=response)
            return await asyncio.wait_for(fut, timeout=timeout)
        finally:
            self._pending_rx = None

    async def start_notify(self, char_uuid: str) -> None:
        client = self._client
        if not client or not client.is_connected:
            raise RuntimeError("Not connected")
        if self._notify_uuid and self._notify_uuid != char_uuid:
            await client.stop_notify(self._notify_uuid)
        if self._notify_uuid == char_uuid:
            return
        await client.start_notify(char_uuid, self._notify_handler)
        self._notify_uuid = char_uuid

    async def stop_notify(self) -> None:
        client = self._client
        if not client or not client.is_connected or not self._notify_uuid:
            self._notify_uuid = None
            return
        uuid = self._notify_uuid
        await client.stop_notify(uuid)
        self._notify_uuid = None


class BleIndustrialApp(ttk.Frame):
    def __init__(self, master: tk.Tk) -> None:
        super().__init__(master, padding=8)
        self.master.title("BLE Industrial Tool")
        self.master.geometry("1380x720")

        self.hub = BleHub()
        self.hub.start()
        self.hub.set_rx_callback(self._on_rx_bytes)

        self.scan_rows: list[tuple[str, Optional[str], int]] = []
        self._scan_rows_all: list[tuple[str, Optional[str], int]] = []

        self.excel_headers: list[str] = [""] * 7
        self.excel_data: list[list[str]] = []
        self._func_col_idx: int = 0
        self._excel_tree: Optional[ttk.Treeview] = None
        self._excel_path: Optional[str] = None
        self._excel_active_sheet: str = ""
        self._excel_sheet_combo: Optional[ttk.Combobox] = None
        self._row_dialog: Optional[tk.Toplevel] = None

        self._build()
        self.master.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build(self) -> None:
        header = ttk.Frame(self)
        header.pack(fill=tk.X, anchor=tk.W, pady=(0, 6))
        ttk.Label(header, text="F. Y. Wu", font=("Segoe UI", 11)).pack(side=tk.LEFT, anchor=tk.W)

        top = ttk.LabelFrame(self, text="Scan", padding=6)
        top.pack(fill=tk.X)

        row = ttk.Frame(top)
        row.pack(fill=tk.X)
        ttk.Label(row, text="Seconds:").pack(side=tk.LEFT)
        self.scan_timeout = tk.DoubleVar(value=5.0)
        ttk.Spinbox(row, from_=1.0, to=120.0, increment=1.0, textvariable=self.scan_timeout, width=6).pack(
            side=tk.LEFT, padx=(4, 12)
        )
        ttk.Button(row, text="Scan", command=self._do_scan).pack(side=tk.LEFT)
        ttk.Label(row, text="Name filter:").pack(side=tk.LEFT, padx=(12, 0))
        self.scan_name_filter = tk.StringVar()
        filter_entry = ttk.Entry(row, textvariable=self.scan_name_filter, width=20)
        filter_entry.pack(side=tk.LEFT, padx=(4, 0))
        filter_entry.bind("<Return>", lambda _e: self._apply_scan_filter())

        mid = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        mid.pack(fill=tk.BOTH, expand=True, pady=6)

        left = ttk.LabelFrame(mid, text="Devices", padding=4)
        mid.add(left)

        self.device_list = tk.Listbox(left, height=10, exportselection=False)
        self.device_list.pack(fill=tk.BOTH, expand=True)
        self.device_list.bind("<Double-Button-1>", self._on_pick_device)

        addr_row = ttk.Frame(left)
        addr_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(addr_row, text="Address:").pack(side=tk.LEFT)
        self.address_var = tk.StringVar()
        ttk.Entry(addr_row, textvariable=self.address_var, width=22).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)

        btn_row = ttk.Frame(left)
        btn_row.pack(fill=tk.X, pady=6)
        ttk.Button(btn_row, text="Connect", command=self._do_connect).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="Disconnect", command=self._do_disconnect).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(btn_row, text="Discover GATT", command=self._do_discover).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(btn_row, text="Auto UUIDs", command=self._do_auto_uuids).pack(side=tk.LEFT, padx=(6, 0))

        opt = ttk.Frame(left)
        opt.pack(fill=tk.X, pady=(4, 0))
        self.auto_uuid_on_connect = tk.BooleanVar(value=True)
        self.auto_subscribe_after_pick = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt,
            text="After connect: auto-pick write / notify UUIDs",
            variable=self.auto_uuid_on_connect,
        ).pack(anchor=tk.W)
        ttk.Checkbutton(
            opt,
            text="After auto-pick: subscribe notify (for RX in log)",
            variable=self.auto_subscribe_after_pick,
        ).pack(anchor=tk.W)

        right = ttk.LabelFrame(mid, text="GATT / traffic", padding=4)
        mid.add(right)

        self.gatt_text = scrolledtext.ScrolledText(right, height=8, wrap=tk.WORD, font=("Consolas", 9))
        self.gatt_text.pack(fill=tk.BOTH, expand=True)

        u = ttk.Frame(right)
        u.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(u, text="Write UUID").grid(row=0, column=0, sticky=tk.W)
        self.write_uuid = tk.StringVar()
        ttk.Entry(u, textvariable=self.write_uuid).grid(row=0, column=1, sticky=tk.EW, padx=4)
        ttk.Label(u, text="Notify UUID").grid(row=1, column=0, sticky=tk.W, pady=(4, 0))
        self.notify_uuid = tk.StringVar()
        ttk.Entry(u, textvariable=self.notify_uuid).grid(row=1, column=1, sticky=tk.EW, padx=4, pady=(4, 0))
        u.columnconfigure(1, weight=1)

        self.write_no_response = tk.BooleanVar(value=False)
        ttk.Checkbutton(right, text="Write without response", variable=self.write_no_response).pack(anchor=tk.W, pady=(4, 0))

        hx = ttk.Frame(right)
        hx.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(hx, text="Hex payload").pack(anchor=tk.W)
        self.hex_var = tk.StringVar()
        ttk.Entry(hx, textvariable=self.hex_var).pack(fill=tk.X, pady=(2, 4))
        br = ttk.Frame(right)
        br.pack(fill=tk.X)
        ttk.Button(br, text="Send", command=self._do_send).pack(side=tk.LEFT)
        ttk.Button(br, text="Subscribe notify", command=self._do_subscribe).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(br, text="Unsubscribe", command=self._do_unsubscribe).pack(side=tk.LEFT, padx=(6, 0))

        excel_fr = ttk.LabelFrame(mid, text="Excel (7 columns)", padding=4)
        mid.add(excel_fr)

        excel_btns = ttk.Frame(excel_fr)
        excel_btns.pack(fill=tk.X)
        ttk.Button(excel_btns, text="Load Excel…", command=self._load_excel_file).pack(side=tk.LEFT)
        ttk.Label(excel_btns, text="Sheet:").pack(side=tk.LEFT, padx=(10, 0))
        self.excel_sheet_var = tk.StringVar()
        self._excel_sheet_combo = ttk.Combobox(
            excel_btns,
            textvariable=self.excel_sheet_var,
            width=24,
            state="readonly",
        )
        self._excel_sheet_combo.pack(side=tk.LEFT, padx=(4, 0))
        self._excel_sheet_combo.bind("<<ComboboxSelected>>", self._on_excel_sheet_changed)
        ttk.Label(
            excel_fr,
            text='Finds cell "address" or "地址" as top-left anchor. Address column = first of 7. '
            "Feature byte from the row where Address=1 (functional attribute column). "
            "Use Sheet to switch worksheets. Double-click a row to read (0x03) or write (0x10). "
            "CRC-16-Modbus on bytes after 55 AA.",
            wraplength=900,
            font=("Segoe UI", 8),
        ).pack(fill=tk.X, anchor=tk.W, pady=(2, 4))

        cols = ("c0", "c1", "c2", "c3", "c4", "c5", "c6")
        tree_wrap = ttk.Frame(excel_fr)
        tree_wrap.pack(fill=tk.BOTH, expand=True)
        yscroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL)
        xscroll = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL)
        self._excel_tree = ttk.Treeview(
            tree_wrap,
            columns=cols,
            show="headings",
            height=12,
            yscrollcommand=yscroll.set,
            xscrollcommand=xscroll.set,
        )
        yscroll.config(command=self._excel_tree.yview)
        xscroll.config(command=self._excel_tree.xview)
        self._excel_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree_wrap.grid_rowconfigure(0, weight=1)
        tree_wrap.grid_columnconfigure(0, weight=1)
        self._excel_tree.bind("<Double-1>", self._on_excel_tree_double_click)
        self._refresh_excel_tree()

        self._mid_pane = mid
        self._pane_layout_applied = False

        def _apply_pane_layout(_event: Optional[tk.Event[Any]] = None) -> None:
            if self._pane_layout_applied:
                return
            w = mid.winfo_width()
            if w < 400:
                return
            left_w = 300
            gatt_w = (900 - 300) // 4  # GATT pane: quarter of original 600px width
            try:
                mid.sashpos(0, left_w)
                mid.sashpos(1, left_w + gatt_w)
                self._pane_layout_applied = True
            except tk.TclError:
                pass

        mid.bind("<Configure>", _apply_pane_layout, add="+")
        self.after(200, _apply_pane_layout)

        logf = ttk.LabelFrame(self, text="Log", padding=4)
        logf.pack(fill=tk.BOTH, expand=True, pady=(6, 0))
        self.log = scrolledtext.ScrolledText(logf, height=10, wrap=tk.WORD, state=tk.DISABLED, font=("Consolas", 9))
        self.log.pack(fill=tk.BOTH, expand=True)

        self.pack(fill=tk.BOTH, expand=True)

    def _log(self, msg: str) -> None:
        self.log.configure(state=tk.NORMAL)
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.configure(state=tk.DISABLED)

    def _on_rx_bytes(self, data: bytes) -> None:
        def go() -> None:
            self._log(f"RX  {fmt_hex(data)}")

        self.after(0, go)

    def _errbox(self, title: str, exc: BaseException) -> None:
        self._log(f"ERROR: {title}: {exc}")
        messagebox.showerror(title, str(exc))

    def _refresh_excel_tree(self) -> None:
        tree = self._excel_tree
        if tree is None:
            return
        tree.delete(*tree.get_children())
        hdrs = (self.excel_headers + [""] * 7)[:7]
        for i in range(7):
            lab = hdrs[i] if hdrs[i].strip() else f"Col{i}"
            tree.heading(f"c{i}", text=lab)
            tree.column(f"c{i}", width=88, stretch=True, anchor=tk.W)
        for i, row7 in enumerate(self.excel_data):
            pad = (row7 + [""] * 7)[:7]
            tree.insert("", tk.END, iid=str(i), values=pad)

    def _apply_excel_sheet_data(
        self,
        path: str,
        sheet_name: str,
        ar: int,
        ac: int,
        headers: list[str],
        data: list[list[str]],
    ) -> None:
        self._func_col_idx = function_attribute_col_index(headers)
        self.excel_headers = (headers + [""] * 7)[:7]
        self.excel_data = [(r + [""] * 7)[:7] for r in data]
        self._excel_active_sheet = sheet_name
        self.excel_sheet_var.set(sheet_name)
        self._refresh_excel_tree()
        self._log(
            f"Excel loaded: {path}  sheet={sheet_name!r}  anchor(row,col)=({ar},{ac})  "
            f"data_rows={len(self.excel_data)}  functional_attribute_col_index={self._func_col_idx}"
        )

    def _load_excel_file(self) -> None:
        path = filedialog.askopenfilename(
            parent=self.master,
            title="Select Excel table",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            sheet_names = excel_sheet_names(path)
            if not sheet_names:
                raise ValueError("Workbook has no worksheets.")
            if self._excel_sheet_combo is not None:
                self._excel_sheet_combo["values"] = sheet_names
            self._excel_path = path
            ar, ac, headers, data, sheet_name = load_excel_seven_from_address_anchor(path)
            self._apply_excel_sheet_data(path, sheet_name, ar, ac, headers, data)
        except Exception as e:
            self._errbox("Excel load", e)

    def _reload_excel_sheet(self, sheet_name: str) -> None:
        if not self._excel_path:
            return
        ar, ac, headers, data, used_sheet = load_excel_seven_from_address_anchor(self._excel_path, sheet_name)
        self._apply_excel_sheet_data(self._excel_path, used_sheet, ar, ac, headers, data)

    def _on_excel_sheet_changed(self, _event: Optional[tk.Event[Any]] = None) -> None:
        if not self._excel_path:
            return
        sheet_name = self.excel_sheet_var.get().strip()
        if not sheet_name or sheet_name == self._excel_active_sheet:
            return
        previous = self._excel_active_sheet
        try:
            self._reload_excel_sheet(sheet_name)
        except Exception as e:
            self.excel_sheet_var.set(previous)
            self._errbox("Excel sheet", e)

    def _on_excel_tree_double_click(self, event: tk.Event[Any]) -> None:
        tree = self._excel_tree
        if tree is None:
            return
        if tree.identify_region(event.x, event.y) != "cell":
            return
        rid = tree.identify_row(event.y)
        if not rid:
            return
        self._open_row_access_dialog(int(rid))

    def _row_values(self, row_index: int) -> tuple[str, ...]:
        if row_index < 0 or row_index >= len(self.excel_data):
            raise IndexError("row index out of range")
        return tuple(str(x) for x in (self.excel_data[row_index] + [""] * 7)[:7])

    def _feature_byte_for_protocol(self) -> int:
        if not self.excel_data:
            raise ValueError("Load an Excel file first.")
        return feature_byte_from_excel_rows(self.excel_data, self._func_col_idx, addr_col_idx=0)

    def _ble_uuids_for_traffic(self) -> tuple[str, str, bool]:
        write_uuid = self.write_uuid.get().strip()
        notify_uuid = self.notify_uuid.get().strip()
        if not write_uuid:
            raise ValueError("Enter write characteristic UUID.")
        if not notify_uuid:
            raise ValueError("Enter notify characteristic UUID.")
        return write_uuid, notify_uuid, not self.write_no_response.get()

    @staticmethod
    def _center_toplevel(win: tk.Toplevel, parent: tk.Misc) -> None:
        win.update_idletasks()
        parent.update_idletasks()
        ww = win.winfo_reqwidth()
        wh = win.winfo_reqheight()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        x = px + max(0, (pw - ww) // 2)
        y = py + max(0, (ph - wh) // 2)
        win.geometry(f"+{x}+{y}")

    def _open_row_access_dialog(self, row_index: int) -> None:
        if not self.excel_data:
            messagebox.showwarning("Excel", "Load an Excel file first.", parent=self.master)
            return
        try:
            row7 = self._row_values(row_index)
        except IndexError:
            return
        addr_s = row7[0].strip()
        if not addr_s:
            messagebox.showwarning("Address", "This row has no address value.", parent=self.master)
            return

        if self._row_dialog is not None and self._row_dialog.winfo_exists():
            self._row_dialog.destroy()

        win = tk.Toplevel(self.master)
        self._row_dialog = win
        win.title(f"Register @ {addr_s}")
        win.transient(self.master)
        win.grab_set()
        win.resizable(False, False)

        frm = ttk.Frame(win, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frm, text=f"Address: {addr_s}").pack(anchor=tk.W)
        val_var = tk.StringVar()
        entry = ttk.Entry(frm, textvariable=val_var, width=36)
        entry.pack(fill=tk.X, pady=(8, 10))
        entry.focus_set()

        btn_row = ttk.Frame(frm)
        btn_row.pack(fill=tk.X)
        read_btn = ttk.Button(btn_row, text="Read", width=10)
        write_btn = ttk.Button(btn_row, text="Write", width=10)
        read_btn.pack(side=tk.LEFT)
        write_btn.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(btn_row, text="Close", width=10, command=win.destroy).pack(side=tk.RIGHT)

        def set_busy(busy: bool) -> None:
            state = tk.DISABLED if busy else tk.NORMAL
            read_btn.configure(state=state)
            write_btn.configure(state=state)

        def on_read() -> None:
            try:
                addr_v = parse_int_for_frame(addr_s) & 0xFFFF
                feat_b = self._feature_byte_for_protocol()
                write_uuid, notify_uuid, response = self._ble_uuids_for_traffic()
                frame = build_read_register_frame(feat_b, addr_v)
            except Exception as e:
                self._errbox("Read", e)
                return
            if not self.hub.connected:
                messagebox.showwarning("BLE", "Not connected.", parent=win)
                return
            set_busy(True)
            self._log(f"Read request @ {addr_s}: {fmt_hex(frame)}")

            def done(f: concurrent.futures.Future[Any]) -> None:
                def ui() -> None:
                    set_busy(False)
                    try:
                        raw = f.result()
                        payload = parse_read_register_response(raw)
                        shown = format_read_data_display(payload)
                        val_var.set(shown)
                        self._log(f"Read OK @ {addr_s}: {shown}  RX {fmt_hex(raw)}")
                    except Exception:
                        val_var.set("fail")
                        self._log(f"Read failed @ {addr_s}")

                self.after(0, ui)

            self.hub.submit(
                self.hub.transact(write_uuid, notify_uuid, frame, response=response),
                on_done=lambda f: done(f),
            )

        def on_write() -> None:
            try:
                addr_v = parse_int_for_frame(addr_s) & 0xFFFF
                data_v = parse_register_value_text(val_var.get())
                feat_b = self._feature_byte_for_protocol()
                write_uuid, notify_uuid, response = self._ble_uuids_for_traffic()
                frame = build_write_register_frame(feat_b, addr_v, data_v)
            except Exception as e:
                self._errbox("Write", e)
                return
            if not self.hub.connected:
                messagebox.showwarning("BLE", "Not connected.", parent=win)
                return
            set_busy(True)
            self._log(f"Write request @ {addr_s}: {fmt_hex(frame)}")

            def done(f: concurrent.futures.Future[Any]) -> None:
                def ui() -> None:
                    set_busy(False)
                    try:
                        f.result()
                        self._log(f"Write OK @ {addr_s}: {val_var.get()}")
                    except Exception as e:
                        self._errbox("Write failed", e)

                self.after(0, ui)

            self.hub.submit(
                self.hub.write(write_uuid, frame, response=response),
                on_done=lambda f: done(f),
            )

        read_btn.configure(command=on_read)
        write_btn.configure(command=on_write)
        win.bind("<Return>", lambda _e: on_write())
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        self._center_toplevel(win, self.master)

    def _future_ui(self, fut: concurrent.futures.Future[Any], ok_msg: Optional[str] = None) -> None:
        try:
            fut.result()
            if ok_msg:
                self._log(ok_msg)
        except Exception as e:
            self._errbox("BLE", e)

    def _device_matches_name_filter(self, name: Optional[str], pattern: str) -> bool:
        needle = pattern.strip()
        if not needle:
            return True
        return needle.lower() in (name or "").lower()

    def _apply_scan_filter(self) -> None:
        if not self._scan_rows_all:
            return
        pattern = self.scan_name_filter.get()
        filtered = [r for r in self._scan_rows_all if self._device_matches_name_filter(r[1], pattern)]
        self.scan_rows = filtered
        self.device_list.delete(0, tk.END)
        for addr, name, rssi in filtered:
            label = f"{addr}  |  {name or '(no name)'}  |  {rssi} dBm"
            self.device_list.insert(tk.END, label)
        if pattern.strip():
            self._log(f"Name filter {pattern!r}: showing {len(filtered)} of {len(self._scan_rows_all)} device(s).")
        else:
            self._log(f"Showing {len(filtered)} device(s).")

    def _do_scan(self) -> None:
        self._log("Scanning...")
        timeout = float(self.scan_timeout.get())

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                rows: list[tuple[str, Optional[str], int]] = f.result()
            except Exception as e:
                self._errbox("Scan failed", e)
                return
            self._scan_rows_all = rows
            self._apply_scan_filter()
            self._log(f"Scan done: {len(rows)} device(s) found.")

        self.hub.submit(self.hub.scan(timeout), on_done=lambda f: self.after(0, lambda: done(f)))

    def _on_pick_device(self, _event: Optional[tk.Event[Any]] = None) -> None:
        sel = self.device_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if 0 <= idx < len(self.scan_rows):
            self.address_var.set(self.scan_rows[idx][0])

    def _do_connect(self) -> None:
        addr = self.address_var.get().strip()
        if not addr:
            messagebox.showwarning("Address", "Enter or select a device address.")
            return
        self._log(f"Connecting to {addr}...")

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                f.result()
            except Exception as e:
                self._errbox("Connect failed", e)
                return
            self._log(f"Connected: {addr}")
            if self.auto_uuid_on_connect.get():
                self._run_auto_uuid_chain(self.auto_subscribe_after_pick.get())

        self.hub.submit(self.hub.connect(addr), on_done=lambda f: self.after(0, lambda: done(f)))

    def _do_disconnect(self) -> None:

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                f.result()
            except Exception as e:
                self._errbox("Disconnect", e)
                return
            self._log("Disconnected.")

        self.hub.submit(self.hub.disconnect(), on_done=lambda f: self.after(0, lambda: done(f)))

    def _do_discover(self) -> None:

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                lines = f.result()
            except Exception as e:
                self._errbox("Discover failed", e)
                return
            self.gatt_text.delete("1.0", tk.END)
            self.gatt_text.insert(tk.END, "\n".join(lines) + "\n")
            self._log("GATT discovery complete.")

        self.hub.submit(self.hub.discover_lines(), on_done=lambda f: self.after(0, lambda: done(f)))

    def _apply_auto_result(self, r: AutoUuidResult) -> None:
        self.write_uuid.set(r.write_uuid or "")
        self.notify_uuid.set(r.notify_uuid or "")
        self.write_no_response.set(r.write_without_response)
        self._log(f"Auto UUIDs: {r.note}")
        if r.write_uuid:
            self._log(f"  Write: {r.write_uuid}")
        if r.notify_uuid:
            self._log(f"  Notify: {r.notify_uuid}")
        self._log(f"  Write without response: {r.write_without_response}")

    def _run_auto_uuid_chain(self, subscribe_after: bool) -> None:
        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                r: AutoUuidResult = f.result()
            except Exception as e:
                self._errbox("Auto UUIDs failed", e)
                return
            self._apply_auto_result(r)
            if subscribe_after and r.notify_uuid:
                self._do_subscribe()

        self.hub.submit(self.hub.auto_pick_uuids(), on_done=lambda f: self.after(0, lambda: done(f)))

    def _do_auto_uuids(self) -> None:
        self._run_auto_uuid_chain(self.auto_subscribe_after_pick.get())

    def _do_send(self) -> None:
        uid = self.write_uuid.get().strip()
        if not uid:
            messagebox.showwarning("UUID", "Enter write characteristic UUID.")
            return
        try:
            payload = parse_hex(self.hex_var.get())
        except ValueError as e:
            messagebox.showwarning("Hex", str(e))
            return
        response = not self.write_no_response.get()

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                f.result()
            except Exception as e:
                self._errbox("Send failed", e)
                return
            self._log(f"TX  {fmt_hex(payload)}")

        self.hub.submit(self.hub.write(uid, payload, response=response), on_done=lambda f: self.after(0, lambda: done(f)))

    def _do_subscribe(self) -> None:
        uid = self.notify_uuid.get().strip()
        if not uid:
            messagebox.showwarning("UUID", "Enter notify characteristic UUID.")
            return

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                f.result()
            except Exception as e:
                self._errbox("Subscribe failed", e)
                return
            self._log(f"Subscribed to notify: {uid}")

        self.hub.submit(self.hub.start_notify(uid), on_done=lambda f: self.after(0, lambda: done(f)))

    def _do_unsubscribe(self) -> None:

        def done(f: concurrent.futures.Future[Any]) -> None:
            try:
                f.result()
            except Exception as e:
                self._errbox("Unsubscribe failed", e)
                return
            self._log("Unsubscribed from notify.")

        self.hub.submit(self.hub.stop_notify(), on_done=lambda f: self.after(0, lambda: done(f)))

    def _on_close(self) -> None:
        try:
            self.hub.stop()
        finally:
            self.master.destroy()


def main() -> None:
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    BleIndustrialApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
